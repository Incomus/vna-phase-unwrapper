from kivy.app import App
from kivy.lang import Builder
from kivy.properties import StringProperty, BooleanProperty
from kivymd.uix.screen import MDScreen
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from pathlib import Path
from datetime import datetime
import traceback
import re
from openpyxl import load_workbook
from xlsxwriter import Workbook
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.axis import ChartLines
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image
import numpy as np
# from scipy.ndimage import gaussian_filter1d
import traceback
import sys
import math
import os
from tkinter import filedialog, Tk


class MainScreen(MDScreen):
    min_jump_default = StringProperty("150")  # Define the default value here
    n_steps_default = StringProperty("1")
    min_offset_default = StringProperty("1")
    max_offset_default = StringProperty("50")
    iterations_default = StringProperty("5")
    scale_reference_path = StringProperty(r'D:\Gleb\S type plot\time_scaling.csv')
    noise_enabled = BooleanProperty(False)
    def on_checkbox_active(self, checkbox, value):
        self.some_option_enabled = value


pd.set_option('display.max_rows', None)
pd.set_option('display.max_columns', None)
    
scale_reference_df = pd.read_csv(r'D:\Gleb\S type plot\time_scaling.csv')

def select_file_dialog(self, title, target_attr, mode="file"):
    root = Tk()
    root.withdraw()

    # Use the current value of the StringProperty as the default path
    default_path = getattr(self, target_attr, "")
    initial_dir = os.path.dirname(default_path) if os.path.isfile(default_path) else default_path or "."
    
    if mode == "folder":
        selected_path = filedialog.askdirectory(
            title=title,
            initialdir=initial_dir
        )
    else:
        selected_path = filedialog.askopenfilename(
            title=title,
            initialdir=initial_dir,
            filetypes=[("All files", "*.*")]
        )
    root.destroy()
    
    if selected_path:
        setattr(self, target_attr, selected_path)

def collect_experiments(main_folder, depth):
    file_list = []
    
    def recursive_search(folder, current_depth, experiment_count):
        # Count Excel files in the current folder
        experiment_count += sum(1 for _ in folder.glob("*.xlsx"))
        for file in folder.glob("*.xlsx"):
            file_list.append(file)
        
        # Check if we can go deeper
        if depth == -1 or current_depth < depth:
            for subfolder in folder.iterdir():
                if subfolder.is_dir():
                    # Recursively search in the subfolder
                    experiment_count = recursive_search(subfolder, current_depth + 1, experiment_count)

        return experiment_count

    # Start the recursive search from the main folder
    experiment_count = recursive_search(main_folder, 0, 0)

    return experiment_count, file_list
    
def summarize_folders():
    while True:
        main_folder_path = input("Input path to the main folder:\n")
        main_folder = Path(main_folder_path)
        if main_folder.exists() and main_folder.is_dir():  # Check if the path exists and is a directory
            break  # Break the loop if the input is valid
        else:
            print(f"'{main_folder_path}' is not a valid directory. Please try again.")
    experiment_count, file_list = collect_experiments(main_folder, 5)
    print(f'Found {experiment_count} experiments')
    input('Press Enter to continue\n')
    return main_folder_path

def choose(text, *options):
    options = list(options)  # Convert the tuple to a list for easier handling
    if isinstance(options, list) and all(isinstance(item, list) for item in options):
        options = options[0]
    while (user_input := input(f'{text}\n({"/".join(options)})\n')) not in options:
        print('Invalid input\n')
    return user_input
    
def recursive_gap_separator(df, n, times, new_column=None, column='final_phase'):
    df = df.copy()
    phase_deriv = [0]
    phase_deriv_abs = [0]
    for i in range(1, len(df)):
        d_phase = df.iloc[i][column] - df.iloc[i - 1][column]
        phase_deriv.append(d_phase)
        phase_deriv_abs.append(abs(d_phase))
    
    df['phase_deriv'] = phase_deriv
    df['group_jump'] = phase_deriv_abs
                                
    cumulative_mask = pd.Series(False, index=df.index)

    for i in range(1, n + 1):
        remaining_mask = ~cumulative_mask

        working_values = df.loc[remaining_mask, 'group_jump'].sort_values().reset_index(drop=True)
        diffs = working_values.diff().fillna(0)

        max_gap_idx = diffs.idxmax()

        separator = working_values[max_gap_idx]
        new_mask = (df['group_jump'] >= separator) & remaining_mask

        # Update cumulative mask
        cumulative_mask = cumulative_mask | new_mask
    unwrapped = [df[column][0]]
    offset = 0
    for i in range(1, len(df)):
        # Detect jump and update offset
        time_val = df.iloc[i]['time']
        in_any_range = any(t0 < time_val < t1 for t0, t1 in zip(times[::2], times[1::2]))
        if (cumulative_mask[i]) and not in_any_range:
            if df.iloc[i]['phase_deriv'] > min_jump:
                offset -= 360
            elif df.iloc[i]['phase_deriv'] < -min_jump:
                offset += 360
        if (cumulative_mask[i]) and in_any_range:
            print('ignored ' + str(time_val))
        if (cumulative_mask[i]) and not in_any_range:
            print('appended ' + str(time_val))
        unwrapped.append(df[column][i] + offset)
    df['final_phase'] = unwrapped
    def mean_of_extremes(series, percent=0.1):
        sorted_vals = series.sort_values()
        n = max(int(len(sorted_vals) * percent), 1)  # At least one value
        bottom_avg = sorted_vals.iloc[:n].mean()
        top_avg = sorted_vals.iloc[-n:].mean()
        return (bottom_avg + top_avg) / 2
    # schizo in
    for t0, t1 in zip(times[::2], times[1::2]):
        if t0 != t1:
            mask_inside = (df['time'] > t0) & (df['time'] < t1)
            
            mask_first_half = (df['time'] > t0) & (df['time'] < (t1+t0)/2)
            first_half = mean_of_extremes(df.loc[mask_first_half, 'final_phase'])
            
            mask_second_half = (df['time'] > (t0+t1)/2) & (df['time'] < t1)
            second_half = mean_of_extremes(df.loc[mask_second_half, 'final_phase'])
            
            # Get the closest data points just before and after the range
            try:
                left_idx = df[df['time'] <= t0].iloc[-1].name
                right_idx = df[df['time'] >= t1].iloc[0].name
            except IndexError:
                # One of the sides has no data
                continue
            
            t_left = df.loc[left_idx, 'time']
            t_right = df.loc[right_idx, 'time']
            p_left = df.loc[left_idx, 'final_phase']
            p_right = df.loc[right_idx, 'final_phase']
            
            # Build linear function over the inside range times
            times_inside = df.loc[mask_inside, 'time']
            if times_inside.empty:
                continue
            
            linear_interp_goal = np.interp(times_inside, [t_left, t_right], [p_left, p_right])
            
            linear_interp_now = np.interp(times_inside, [t_left, t_right], [first_half, second_half])

            # Adjust original values so their average matches the linear function
            original_phases = df.loc[mask_inside, 'final_phase']
            avg_difference = linear_interp_goal - linear_interp_now
            corrected_phases = original_phases + avg_difference

            df.loc[mask_inside, 'final_phase'] = corrected_phases   
    # schizo out
    
    if new_column != None:
        df[new_column] = df['final_phase']
    return df
    
def recursive_crap_separator(df, n, times, new_column=None, column='final_phase'):
    df = df.copy()
    phase_deriv = [0]
    phase_deriv_abs = [0]
    for i in range(1, len(df)):
        d_phase = df[column][i] - df[column][i - 1]
        phase_deriv.append(d_phase)
        phase_deriv_abs.append(abs(d_phase))
    
    df['phase_deriv'] = phase_deriv
    df['group_jump'] = phase_deriv_abs
                                
    cumulative_mask = pd.Series(False, index=df.index)

    for i in range(1, n + 1):
        remaining_mask = ~cumulative_mask

        working_values = df.loc[remaining_mask, 'group_jump'].sort_values().reset_index(drop=True)
        diffs = working_values.diff().fillna(0)

        max_gap_idx = diffs.idxmax()

        separator = working_values[max_gap_idx]
        new_mask = (df['group_jump'] >= separator) & remaining_mask

        # Update cumulative mask
        cumulative_mask = cumulative_mask | new_mask
    unwrapped = list(df.iloc[:6][column])
    offset = 0
    valids = []
    interval = 4
    for i in range(6, len(df)):
        # Detect jump and update offset
        time_val = df.iloc[i]['time']
        in_any_range = any(t0 < time_val < t1 for t0, t1 in zip(times[::2], times[1::2]))
        if (cumulative_mask[i]) and not in_any_range:
            current_val = df.iloc[i]['phase_deriv']
            past_vals = df.iloc[i-interval:i]['phase_deriv'].values[::-1]
            greater_than_08 = [j for j, val in enumerate(past_vals) if abs(val) > 0.8 * abs(current_val)]
            greater_than_neg = [j for j, val in enumerate(past_vals) if val * current_val < 0]
            if i < len(df)-interval-1:
                future_vals = df.iloc[i:i+interval]['phase_deriv'].values[::-1]
                greater_08_future = [j for j, val in enumerate(future_vals) if abs(val) > 0.8 * abs(current_val)]
                greater_neg_future = [j for j, val in enumerate(future_vals) if val * current_val < 0]
                if not (greater_than_08 and greater_than_neg and min(greater_than_08) in greater_than_neg):
                    if not (greater_08_future and greater_neg_future and min(greater_08_future) in greater_neg_future):
                        if df['phase_deriv'][i] > min_jump:
                            offset -= 360
                            valids.append(i)
                        elif df['phase_deriv'][i] < -min_jump:
                            offset += 360
                            valids.append(i)
            if greater_than_08 and greater_than_neg and min(greater_than_08) in greater_than_neg:
                if i-min(greater_than_08)-1 in valids:
                    pass
                elif df.iloc[i]['phase_deriv'] > min_jump:
                    offset -= 360
                    valids.append(i)
                elif df.iloc[i]['phase_deriv'] < -min_jump:
                    offset += 360
                    valids.append(i)
        unwrapped.append(df[column][i] + offset)
    df['final_phase'] = unwrapped
    
    # schizo in
    for t0, t1 in zip(times[::2], times[1::2]):
        mask_inside = (df['time'] > t0) & (df['time'] < t1)
        
        # Get the closest data points just before and after the range
        try:
            left_idx = df[df['time'] <= t0].iloc[-1].name
            right_idx = df[df['time'] >= t1].iloc[0].name
        except IndexError:
            # One of the sides has no data
            continue
        
        t_left = df.loc[left_idx, 'time']
        t_right = df.loc[right_idx, 'time']
        p_left = df.loc[left_idx, 'final_phase']
        p_right = df.loc[right_idx, 'final_phase']
        
        # Build linear function over the inside range times
        times_inside = df.loc[mask_inside, 'time']
        if times_inside.empty:
            continue
        
        linear_interp = np.interp(times_inside, [t_left, t_right], [p_left, p_right])

        # Adjust original values so their average matches the linear function
        original_phases = df.loc[mask_inside, 'final_phase']
        print(linear_interp)
        print(original_phases)
        avg_difference = linear_interp - original_phases.mean()
        corrected_phases = original_phases + avg_difference

        df.loc[mask_inside, 'final_phase'] = corrected_phases   
    # schizo out
    
    if new_column != None:
        df[new_column] = df['final_phase']
    return df
    
def recursive_point_separator(df, n, min_offset, max_offset, times, new_column=None, column='final_phase'):
    df = df.copy()
    offset_addit = -1
    offset = min_offset
    for iter in range(1, n + 1):
        unwrapped = list(df[column].iloc[:offset])
        phase_plus_360 = df[column] + 360
        phase_minus_360 = df[column] - 360
        
        for i in range(offset, len(df)-offset):
            neighbors = [df[column].iloc[j] for j in range(i - offset, i + offset + 1) if j != i]
            avg_neighbors = sum(neighbors) / len(neighbors)
            phase_plus_dif = abs(avg_neighbors - phase_plus_360[i]) < abs(avg_neighbors - df[column][i])
            phase_minus_dif = abs(avg_neighbors - phase_minus_360[i]) < abs(avg_neighbors - df[column][i])
            time_val = df.iloc[i]['time']
            in_any_range = any(t0 < time_val < t1 for t0, t1 in zip(times[::2], times[1::2]))
            if phase_plus_dif and not in_any_range:
                unwrapped.append(phase_plus_360[i])
                print(iter, '+360', f'offset:{offset}', round(df['time'].iloc[i], 6), round(abs(avg_neighbors - phase_plus_360[i]), 3))
            elif phase_minus_dif and not in_any_range:
                unwrapped.append(phase_minus_360[i])
                print(iter, '-360', f'offset:{offset}', round(df['time'].iloc[i], 6), round(abs(avg_neighbors - phase_minus_360[i]), 3))
            else:
                unwrapped.append(df[column][i])
        unwrapped.extend(df[column].iloc[-offset:])
        df[column] = unwrapped
        if new_column != None:
            df[new_column] = unwrapped
        offset_addit *= -1
        if offset_addit < 0:
            offset = min_offset
            max_offset = int(min_offset + (max_offset - min_offset)/3)
        else:
            offset = max_offset
    return df
    
def read_file(file):
    df = pd.read_excel(file)
    
    df = pd.DataFrame({'phase': df.iloc[:, 2],
                                'time': df.iloc[:, 1],
                                'sweep_time': df.iloc[-1, 1],
                                'file_name': file.stem,
                                'final_phase': df.iloc[:, 2]})
    return df

def scale_dataframe(df, reference):
    # Extracting n_points and t_vna values from the reference
    reference = reference.set_index(["n_points", "t_vna"])
    
    # Checking if df length matches a reference n_points
    if len(df) in reference.index.get_level_values("n_points"):
        matched_n_points = len(df)
        max_x = df["time"].max()
        
        # Finding matching t_vna within +-0.2 range
        possible_t_vna = reference.loc[matched_n_points].index
        matched_t_vna = [t for t in possible_t_vna if abs(max_x - t) <= 0.2]
        print(matched_t_vna)
        if matched_t_vna:
            matched_t_vna = min(matched_t_vna, key=lambda t: abs(max_x - t))  # Take the closest match
            scale_factors = reference.loc[(matched_n_points, matched_t_vna)].values.flatten()
            
            # Applying the scaling factor to df["y"]
            df["time"] *= scale_factors[0]  # Assuming we use the first column's scale factor
    if matched_t_vna:
        return df, scale_factors[0]
    else:
        return df, None

def process_files(main_folder_path, setup):
    main_folder = Path(main_folder_path)
    experiment_count, file_list = collect_experiments(main_folder, 5)
    for file in file_list:
        try:
            sheet_name = 'Sheet1'
            df = read_file(file)
            times = []
            print("")
            print(file)
            if setup == 'y':
                repeats = int(input('Number of noise zones:\n'))
                while repeats > 0:
                    time0 = float(input('Choose time where noise starts (s):\n'))
                    times.append(time0)
                    time1 = float(input('Choose time where noise ends (s):\n'))
                    times.append(time1)
                    repeats += -1
            else:
                time0 = df.iloc[0]['time'] - 1
                times.append(time0)
                time1 = time0
                times.append(time1)
            if override == '1':
                df = recursive_crap_separator(df, 1, times, 'iter_1')
            else:
                df = recursive_gap_separator(df, 1, times, 'iter_1')
            if n_steps != '1':
                df = recursive_point_separator(df, iterations, min_offset, max_offset, times, 'iter_2')
            if n_steps == '3':
                df = recursive_gap_separator(df, 1, times, 'iter_3')
            
            df, scale_factor = scale_dataframe(df, scale_reference_df)
            if scale_factor is not None:
                times = [x * scale_factor for x in times]
            
            steps = ['iter_1', 'iter_2', 'iter_3']
            workbook = load_workbook(file)
            sheet = workbook[sheet_name]
            
            # Write raw phase
            col_num = 3
            for row, val in enumerate(df['phase'], start=2):
                sheet.cell(row=1, column=col_num, value='Phase raw')
                sheet.cell(row=row, column=col_num, value=val)
            
            # Write new time
            col_num += 1
            for row, val in enumerate(df['time'], start=2):
                sheet.cell(row=1, column=col_num, value='Scaled time')
                sheet.cell(row=row, column=col_num, value=val)
            
            # Write phase iterations
            for idx, col in enumerate(steps):
                if (col == 'iter_2' and n_steps == '1') or (col == 'iter_3' and n_steps != '3'):
                    continue
                phase_col = df[col]
                header = f'Phase iteration {idx+1}'
                col_num += 1
                sheet.cell(row=1, column=col_num, value=header)
                for row, val in enumerate(phase_col, start=2):
                    sheet.cell(row=row, column=col_num, value=val)
            
            # Write time and jump values
            # extra_cols = {
                # 'Jump time': sorted_phase['time'],
                # 'Jump size': sorted_phase['group_jump'],
            # }

            # for header, col_data in extra_cols.items():
                # col_num = col_num + 1
                # sheet.cell(row=1, column=col_num, value=header)
                # for row, val in enumerate(col_data, start=2):
                    # sheet.cell(row=row, column=col_num, value=val)
            
            col_num += 1
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=col_num, max_col=col_num):
                for cell in row:
                    cell.value = None
            if setup == 'y':
                # Create a dummy series for shading the ignored region
                ignore_y = []

                for i in range(len(df)):
                    time_val = df.iloc[i]['time']
                    in_any_range = any(t0 < time_val < t1 for t0, t1 in zip(times[::2], times[1::2]))
                    if in_any_range:
                        ignore_y.append(df.iloc[i]['iter_1'])
                    else:
                        ignore_y.append(None)

                sheet.cell(row=1, column=col_num, value='Ignore')
                for row, val in enumerate(ignore_y, start=2):
                    sheet.cell(row=row, column=col_num, value=val)
            
            max_row = sheet.max_row
            max_col = sheet.max_column
            data = []
            for row in sheet.iter_rows(min_row=1, max_row=max_row, max_col=max_col, values_only=True):
                data.append(row)
            workbook.save(file)
            workbook.close()
            
            # x_vals = [sheet.cell(row=i, column=2).value for i in range(2, 2 + len(phase_degrees_np_1))]
            # y_vals = list(phase_degrees_np_1)
            wb = Workbook(file)
            ws = wb.add_worksheet(sheet_name)
            for row_idx, row in enumerate(data):
                for col_idx, val in enumerate(row):
                    if isinstance(val, (int, float)):
                        ws.write_number(row_idx, col_idx, val)
                    elif val != None:
                        ws.write_string(row_idx, col_idx, str(val))
            # ws.write('A1', 'Time[s]')
            # ws.write('B1', 'Phase')
            # ws.write_column('A2', x_vals)
            # ws.write_column('B2', y_vals)
            chart = wb.add_chart({'type': 'scatter', 'subtype': 'straight'})
            
            col_num = 2
            chart.add_series({
                'name':       'Raw phase data',
                'categories': [sheet_name, 1, 3, len(df), 3],  # X values
                'values':     [sheet_name, 1, col_num, len(df), col_num],  # Y values
                'marker':     {'type': 'none'}
            })
            chart.set_title({'name': 'Raw vs Processed Data'})
            for idx, col in enumerate(steps):
                if (col == 'iter_2' and n_steps == '1') or (col == 'iter_3' and n_steps != '3'):
                    continue
                col_num = 4 + idx
                chart.add_series({
                    'name':       f'Processed phase data iteration {idx+1}',
                    'categories': [sheet_name, 1, 3, len(df), 3],  # X values
                    'values':     [sheet_name, 1, col_num, len(df), col_num],  # Y values
                    'marker':     {'type': 'none'}
                })
            if setup == 'y':
                col_num += 1
                chart.add_series({
                    'name':       'Ignored Interval',
                    'categories': [sheet_name, 1, 3, len(df), 3],
                    'values':     [sheet_name, 1, col_num, len(df), col_num],
                    'marker':     {'type': 'none'},
                    'line':       {'color': '#CCCCCC', 'width': 3, 'dash_type': 'square_dot'},
                })
            # print(max(phase['time']), max(phase_degrees_np_1))
            # max_y = max(max(phase['time']), max(phase_degrees_np_1))
            # min_y = min(min(phase['time']), min(phase_degrees_np_1))
            chart.set_x_axis({
                'name': 'Time, s', 
                'min': 0, 
                'max': round(max(df['time'])), 
                'major_gridlines': {'visible': True}})
            chart.set_y_axis({'name': 'Phase, deg', 
                # 'min': math.floor(min_y) - 20, 
                # 'max': math.ceil(max_y) + 20, 
                'major_gridlines': {'visible': True}})
            chart.set_legend({'position': 'bottom'})
            chart.set_size({'width': 1400, 'height': 600})
            ws.insert_chart('F3', chart)
            wb.close()
            print(file.stem, ' passed')
            

            
            
            
        except Exception as e:
            tb = traceback.extract_tb(sys.exc_info()[2])[0]
            print(f"Error in '{file.stem}', line {tb.lineno}: {e}")
            try:
                tb2 = traceback.extract_tb(sys.exc_info()[2])[1]
                print(f"Line 2: {tb2.lineno}")
            except:
                None
    
def get_gap(phase, phase_deriv, gaps_df):
    for i in range(1, len(phase)):
        # Detect jump and update gaps
        if (
            phase_deriv['group_high'][i]
        ):
            temp_df = pd.DataFrame({
                    'file_name': phase_deriv['file_name'][i],
                    'n_points': len(phase),
                    'sweep_time': phase_deriv['sweep_time'][0],
                    'time_of_gap': phase_deriv['time'][i],
                    'gap': phase_deriv['phase_deriv'][i]
                    }, index=[0])
            gaps_df = pd.concat([gaps_df, temp_df], ignore_index=True)

    return gaps_df
   
def get_gaps(main_folder_path):
    wb = Workbook()
    main_folder = Path(main_folder_path)
    experiment_count, file_list = collect_experiments(main_folder, 5)
    ws = wb.create_sheet(title='gaps')
    gaps_df = pd.DataFrame()
    for file in file_list:
        try:
            phase, phase_deriv = phase_group(file)
            gaps_df = get_gap(phase, phase_deriv, gaps_df)
            print(file, ' passed')
        except Exception as e:
            tb = traceback.extract_tb(sys.exc_info()[2])[0]
            tb2 = traceback.extract_tb(sys.exc_info()[2])[1]
            print(f"Error in '{file.stem}', line {tb.lineno}, to {tb2.lineno}: {e}")
    
    for r in dataframe_to_rows(gaps_df, index=False, header=True):
        ws.append(r)
    file_name = f"{input(f'Input excel file name\n')}.xlsx"
    wb.save(file_name)
# Main execution
if __name__ == "__main__":
    
    try:
        min_jump = 150
        n_steps = '1'
        override = '0'
        min_offset = 1
        max_offset = 50
        iterations = 5
        setup = 1
        user_input = '0'
        while user_input != '4':
            os.system('cls')
            if setup > 0:
                text = 'not '
            else:
                text = ''
                
            user_input = choose(f'Press a number to:\n1: Change default settings (min jump={min_jump}, number of steps (legacy setting)={n_steps})\n\
2: Noise ignore (time interval for each file) will {text}be asked\n\
3: Change default step 2 settings (legacy settings)\n\
4: Continue', '1', '2', '3', '4')
            if user_input == '1':
                min_jump = int(input('Set minimum jump\n'))
                n_steps = choose('Set number of algorithm steps (default: 1)', '1', '2', '3')
                override = choose('Set override (default: 0)', '0', '1')
            if user_input == '2':
                setup *= -1
            if user_input == '3':
                min_offset = int(input('Set minimum offset\n'))
                max_offset = int(input('Set maximum offset\n'))
                iterations = int(input('Set iterations\n'))
                
        if setup > 0:
            setup = 'n'
        else:
            setup = 'y'
        main_folder_path = summarize_folders()

        process_files(main_folder_path, setup)

    except Exception as e:
        print(f"Error: {e}\nType: {type(e).__name__}\n")
        traceback.print_exc()
    input("Press Enter to exit\n")
